"""
EU skills-shortage data provider.

Downloads and parses the Cedefop Labour and Skills Shortage Index (CLSSI) — the
official EU dataset of shortage scores by occupation and country — and returns
the top shortage occupations for the countries we feature in the report.

The dataset is a public Excel file (one sheet per country plus an EU27 sheet).
Each row has an occupation group and a "Labour Shortage Index" from 1 (low) to
4 (severe). We surface the occupations at or above a configurable threshold.

Everything here is best-effort: if the download or parse fails, we return an
empty result and the report falls back to news-based skills signals.
"""

from __future__ import annotations

import io

import config
from radar.utils.http import get
from radar.utils.logging import get_logger

logger = get_logger(__name__)

# Column positions in each CLSSI sheet (0-indexed), based on the header row:
#   0: Main Occupation Group   1: Occupation Group (2 digit)
#   2: Labour Shortage Index   3: LSI (Comp.)   4..6: LSI1..LSI3
_OCC_COL = 1
_INDEX_COL = 2


def _download_workbook():
    """Fetch the CLSSI Excel file and open it with openpyxl. None on failure."""
    resp = get(config.CEDEFOP_CLSSI_URL)
    if resp is None:
        return None
    try:
        import openpyxl  # imported lazily so the project runs without it
    except ImportError:
        logger.warning("openpyxl not installed — skipping Cedefop skills data.")
        return None
    try:
        return openpyxl.load_workbook(
            io.BytesIO(resp.content), read_only=True, data_only=True
        )
    except Exception as exc:  # corrupt download, etc.
        logger.warning("Could not open Cedefop workbook: %s", exc)
        return None


def _top_shortages(workbook, sheet_code: str, limit: int) -> list[tuple[str, float]]:
    """Return [(occupation, index), ...] above threshold for one sheet."""
    if sheet_code not in workbook.sheetnames:
        return []
    ws = workbook[sheet_code]
    found: list[tuple[str, float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= _INDEX_COL:
            continue
        occupation = row[_OCC_COL]
        index = row[_INDEX_COL]
        if not occupation or not isinstance(index, (int, float)):
            continue
        if float(index) >= config.CEDEFOP_SHORTAGE_THRESHOLD:
            found.append((str(occupation).strip(), round(float(index), 1)))
    # Highest shortage first; keep the strongest signals.
    found.sort(key=lambda pair: pair[1], reverse=True)
    return found[:limit]


def get_shortage_data(limit_per_country: int = 5) -> dict[str, list[tuple[str, float]]]:
    """
    Return a mapping of display name -> top shortage occupations.

    Example: {"EU27": [("Health professionals", 4.0), ...], "Sweden": [...]}
    Returns {} if the dataset cannot be fetched/parsed (report degrades safely).
    """
    workbook = _download_workbook()
    if workbook is None:
        return {}

    data: dict[str, list[tuple[str, float]]] = {}
    try:
        for display_name, sheet_code in config.CEDEFOP_FEATURED.items():
            shortages = _top_shortages(workbook, sheet_code, limit_per_country)
            if shortages:
                data[display_name] = shortages
    finally:
        workbook.close()

    logger.info("Cedefop skills data: %d country sheet(s) with shortages", len(data))
    return data
